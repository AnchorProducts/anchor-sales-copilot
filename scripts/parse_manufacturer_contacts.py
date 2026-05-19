#!/usr/bin/env python3
"""
Parse "Manufacturer Rep Lists" Excel into a clean JSON of contacts.

Run:
  python3 scripts/parse_manufacturer_contacts.py <path-to-xlsx> > scripts/manufacturer_contacts.json

Output schema (per row):
  manufacturer: str
  first_name: str | None
  last_name: str | None
  full_name: str | None
  email: str | None        # lowercased
  phone: str | None
  cell: str | None
  title: str | None
  territory: str | None
  region: str | None       # only set for the Sika regional grid
  raw: dict                # the source row for debugging
"""

import json
import re
import sys
from pathlib import Path

import openpyxl


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s


def clean_phone(v):
    s = clean(v)
    if not s:
        return None
    # Leave the original formatting alone; just strip whitespace.
    return s


def clean_email(v, fallback_domain=None):
    s = clean(v)
    if not s:
        return None
    s = s.lower()
    if "@" not in s and fallback_domain:
        s = f"{s}@{fallback_domain}"
    if "@" not in s:
        return None
    # Drop obvious garbage
    if not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", s):
        return None
    return s


def make_full(first, last):
    f = clean(first)
    l = clean(last)
    if f and l:
        return f"{f} {l}"
    return f or l


def is_data_row_name(v):
    """Skip section headers that appear in the name columns of the Sika grid."""
    if not v:
        return False
    s = str(v).strip().upper()
    return s not in {"SALES", "TECHNICAL", "PHONE", "EMAIL", "EXT", "OFFICE", "EXT/OFFICE", "FUNCTION · TERRITORY"}


# ── Per-sheet parsers ────────────────────────────────────────────────────────

def parse_duro_last(ws):
    """Duro-Last: Lastname Firstname Rep# Status Territory Address1 Address2 City1 State Zip1 Phone Fax Cell Email"""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    for row in rows[1:]:
        last, first, _rep, _status, territory, _a1, _a2, _city, _state, _zip, phone, _fax, cell, email = (row + (None,) * 14)[:14]
        if not (last or first or email):
            continue
        out.append({
            "manufacturer": "Duro-Last",
            "first_name": clean(first),
            "last_name": clean(last),
            "full_name": make_full(first, last),
            "email": clean_email(email),
            "phone": clean_phone(phone),
            "cell": clean_phone(cell),
            "title": None,
            "territory": clean(territory),
            "region": None,
            "raw": {"phone": phone, "cell": cell, "email": email, "territory": territory},
        })
    return out


def parse_fibertite(ws):
    """FiberTite: Name | Email Address"""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        name, email = (row + (None, None))[:2]
        name = clean(name)
        email = clean_email(email)
        if not name and not email:
            continue
        first = last = None
        if name:
            parts = name.split()
            if len(parts) == 1:
                first = parts[0]
            else:
                first = parts[0]
                last = " ".join(parts[1:])
        out.append({
            "manufacturer": "FiberTite",
            "first_name": first,
            "last_name": last,
            "full_name": name,
            "email": email,
            "phone": None,
            "cell": None,
            "title": None,
            "territory": None,
            "region": None,
            "raw": {"name": name, "email": email},
        })
    return out


def parse_sika_technical(ws):
    """Sika (Technical): First, Last, Role, Region/Division, Email"""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        first, last, role, region, email = (row + (None,) * 5)[:5]
        if not (first or last or email):
            continue
        out.append({
            "manufacturer": "Sika",
            "first_name": clean(first),
            "last_name": clean(last),
            "full_name": make_full(first, last),
            "email": clean_email(email, "us.sika.com"),
            "phone": None,
            "cell": None,
            "title": clean(role),
            "territory": clean(region),
            "region": clean(region),
            "raw": {"role": role, "region": region},
        })
    return out


# Schemas for the Sika regional grid. Column offsets per region after
# manually inspecting row 3 of the sheet.
SIKA_REGIONS = [
    {"start": 0,  "name": "Mid-Atlantic",  "states": "DC, DE, MD, NJ, PA, VA, WV",          "cols": {"last": 1,  "first": 2,  "phone": 3,  "email": 4,  "function": 6,  "sales_num": 7}},
    {"start": 9,  "name": "Northeast",     "states": "CT, MA, ME, NH, NY, RI, VT",          "cols": {"last": 10, "first": 11, "phone": 12, "email": 13, "function": 15, "sales_num": 16}},
    {"start": 18, "name": "Southeast",     "states": "AL, FL, GA, MS, NC, SC, TN",          "cols": {"last": 19, "first": 20, "phone": 21, "email": 22, "function": 23, "sales_num": 24}},
    {"start": 26, "name": "Midwest East",  "states": "IN, KY, MI, OH",                       "cols": {"last": 27, "first": 28, "phone": 29, "email": 30, "function": 31, "sales_num": 32}},
    {"start": 34, "name": "Midwest West",  "states": "IA, IL, KS, MN, MO, ND, NE, SD, WI",   "cols": {"last": 35, "first": 36, "phone": 37, "email": 38, "function": 40, "sales_num": 41}},
    {"start": 43, "name": "Southwest",     "states": "AR, LA, OK, TX",                       "cols": {"last": 44, "first": 45, "phone": 46, "email": 47, "function": 49, "sales_num": 50}},
    {"start": 52, "name": "Mountain",      "states": "CO, ID, MT, NM, UT, WY",              "cols": {"last": 53, "first": 54, "phone": 55, "email": 56, "function": 58, "sales_num": 59}},
    {"start": 61, "name": "West North",    "states": "AK, CA(N), HI, NV, OR, WA",            "cols": {"last": 62, "first": 63, "phone": 64, "email": 65, "function": 67, "sales_num": 68}},
    {"start": 70, "name": "West South",    "states": "AZ, CA(S)",                            "cols": {"last": 71, "first": 72, "phone": 73, "email": 74, "function": 76, "sales_num": 77}},
]


def parse_sika_main(ws):
    """Sika main: wide grid with 9 regions side-by-side."""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    # Data rows start at index 3 (row 4 in spreadsheet).
    for row in rows[3:]:
        for region in SIKA_REGIONS:
            cols = region["cols"]
            last_val = row[cols["last"]] if cols["last"] < len(row) else None
            first_val = row[cols["first"]] if cols["first"] < len(row) else None
            email_val = row[cols["email"]] if cols["email"] < len(row) else None

            # Skip section headers ("SALES", "TECHNICAL") that appear in name columns.
            if not is_data_row_name(last_val) or not is_data_row_name(first_val):
                continue
            if not (clean(last_val) or clean(first_val)):
                continue

            phone = row[cols["phone"]] if cols["phone"] < len(row) else None
            function = row[cols["function"]] if cols["function"] < len(row) else None

            out.append({
                "manufacturer": "Sika",
                "first_name": clean(first_val),
                "last_name": clean(last_val),
                "full_name": make_full(first_val, last_val),
                "email": clean_email(email_val, "us.sika.com"),
                "phone": clean_phone(phone),
                "cell": None,
                "title": clean(function),
                "territory": region["states"],
                "region": region["name"],
                "raw": {"sales_num": row[cols["sales_num"]] if cols["sales_num"] < len(row) else None},
            })
    return out


def parse_standard_contact_sheet(manufacturer, ws):
    """For sheets with: First Name, Last Name, Phone Number, Cell, Title, Email"""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        first, last, phone, cell, title, email = (row + (None,) * 6)[:6]
        if not (first or last or email):
            continue
        out.append({
            "manufacturer": manufacturer,
            "first_name": clean(first),
            "last_name": clean(last),
            "full_name": make_full(first, last),
            "email": clean_email(email),
            "phone": clean_phone(phone),
            "cell": clean_phone(cell),
            "title": clean(title),
            "territory": None,
            "region": None,
            "raw": {},
        })
    return out


def parse_gaf(ws):
    """GAF: City, GAF Employee, Position, Email, Phone — City fills down."""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    current_city = None
    for row in rows[1:]:
        city, name, position, email, phone = (row + (None,) * 5)[:5]
        if clean(city):
            current_city = clean(city)
        name = clean(name)
        if not name and not clean_email(email):
            continue
        first = last = None
        if name:
            parts = name.split()
            if len(parts) == 1:
                first = parts[0]
            else:
                first = parts[0]
                last = " ".join(parts[1:])
        out.append({
            "manufacturer": "GAF",
            "first_name": first,
            "last_name": last,
            "full_name": name,
            "email": clean_email(email),
            "phone": clean_phone(phone),
            "cell": None,
            "title": clean(position),
            "territory": current_city,
            "region": current_city,
            "raw": {"city": current_city},
        })
    return out


def main(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_rows = []

    if "Duro-Last" in wb.sheetnames:
        all_rows.extend(parse_duro_last(wb["Duro-Last"]))
    if "FiberTite" in wb.sheetnames:
        all_rows.extend(parse_fibertite(wb["FiberTite"]))
    if "Sika (Technical)" in wb.sheetnames:
        all_rows.extend(parse_sika_technical(wb["Sika (Technical)"]))
    if "Sika" in wb.sheetnames:
        all_rows.extend(parse_sika_main(wb["Sika"]))
    for std in ("Panelclaw", "Aerocompact", "Unirac", "Pegasus"):
        if std in wb.sheetnames:
            all_rows.extend(parse_standard_contact_sheet(std, wb[std]))
    if "GAF" in wb.sheetnames:
        all_rows.extend(parse_gaf(wb["GAF"]))

    # De-dupe on (manufacturer, lower(email)) when email exists. Email-less
    # rows are kept as separate entries.
    seen = set()
    out = []
    for r in all_rows:
        key = (r["manufacturer"], (r["email"] or "").lower()) if r["email"] else None
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        out.append(r)

    json.dump(out, sys.stdout, indent=2)
    sys.stdout.write("\n")
    print(f"# {len(out)} contacts across "
          f"{len({r['manufacturer'] for r in out})} manufacturers",
          file=sys.stderr)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: parse_manufacturer_contacts.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(2)
    main(Path(sys.argv[1]))
